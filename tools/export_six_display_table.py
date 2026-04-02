from __future__ import annotations

import csv
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SRC_DIR = Path(r"C:\Users\clearlove\Desktop\six")
FEATURE_DIR = Path(
    r"D:\CNTDATA\CNTA_ML_Project\reports\arbitrary_50000_bend_ranking_20260401_193842_tiered\features"
)
OUT_XLSX = SRC_DIR / "six_display_table.xlsx"
OUT_CSV = SRC_DIR / "six_display_table.csv"


FILENAME_PATTERN = re.compile(
    r"^(?P<sample>No\d+)\s+"
    r"(?P<al2o3_power>\d+)w\s+(?P<al2o3_thickness>[\d.]+)nm\s+"
    r"(?P<fe_power>\d+)w\s+(?P<fe_thickness>[\d.]+)nm\s+"
    r"(?P<ar>\d+)\s+(?P<h2>\d+)\s+(?P<c2h4>\d+)\s+"
    r"(?P<anneal_temp>\d+)\s+(?P<growth_temp>\d+)\s+"
    r"(?P<anneal_time>\d+)min\s+(?P<growth_time>\d+)min\s+"
    r"(?P<tail>.+?)$"
)


HEADERS = [
    "序号",
    "图片文件名",
    "样品编号",
    "Al2O3功率(W)",
    "Al2O3厚度(nm)",
    "Fe功率(W)",
    "Fe厚度(nm)",
    "氩气流量",
    "氢气流量",
    "乙烯流量",
    "退火温度(℃)",
    "生长温度(℃)",
    "退火时间(min)",
    "生长时间(min)",
    "位置/备注",
    "放大倍数",
    "迂曲度",
    "波曲度",
    "平均直径(nm)",
    "取向度",
    "覆盖密度(%)",
    "曲率",
]


def build_rows() -> list[list[object]]:
    rows: list[list[object]] = []
    for img_path in sorted(SRC_DIR.glob("*.png")):
        parts = img_path.stem.split("-", 2)
        seq = int(parts[0]) if len(parts) >= 1 else None
        base_name = parts[2] + ".png" if len(parts) >= 3 else img_path.name

        feature_path = FEATURE_DIR / f"{Path(base_name).stem}__features.json"
        if not feature_path.exists():
            raise FileNotFoundError(f"Missing feature file for {img_path.name}: {feature_path}")

        obj = json.loads(feature_path.read_text(encoding="utf-8"))
        feats = obj["features"]

        match = FILENAME_PATTERN.match(Path(base_name).stem)
        if not match:
            raise ValueError(f"Cannot parse process params from: {base_name}")
        gd = match.groupdict()

        rows.append(
            [
                seq,
                img_path.name,
                gd["sample"],
                int(gd["al2o3_power"]),
                float(gd["al2o3_thickness"]),
                int(gd["fe_power"]),
                float(gd["fe_thickness"]),
                int(gd["ar"]),
                int(gd["h2"]),
                int(gd["c2h4"]),
                int(gd["anneal_temp"]),
                int(gd["growth_temp"]),
                int(gd["anneal_time"]),
                int(gd["growth_time"]),
                gd["tail"],
                50000,
                float(feats["tortuosity_v2"]) if feats.get("tortuosity_v2") is not None else None,
                float(feats["waviness_ratio_v2"]) if feats.get("waviness_ratio_v2") is not None else None,
                float(feats["diameter"]) if feats.get("diameter") is not None else None,
                float(feats["alignment"]) if feats.get("alignment") is not None else None,
                float(feats["density"]) if feats.get("density") is not None else None,
                float(feats["curvature_nm_v3_trimmed_mean_length"])
                if feats.get("curvature_nm_v3_trimmed_mean_length") is not None
                else None,
            ]
        )
    rows.sort(key=lambda row: int(row[0]))
    return rows


def write_csv(rows: list[list[object]]) -> None:
    with OUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(HEADERS)
        writer.writerows(rows)


def write_xlsx(rows: list[list[object]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "六张图参数表"
    ws.append(HEADERS)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            if cell.column <= 16:
                cell.alignment = center

    num_formats = {
        5: "0.0",
        7: "0.0",
        17: "0.000",
        18: "0.0000",
        19: "0.00",
        20: "0.0000",
        21: "0.00",
        22: "0.000000",
    }
    for col_idx, fmt in num_formats.items():
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row_idx, col_idx).number_format = fmt

    widths = {
        1: 8,
        2: 55,
        3: 12,
        4: 14,
        5: 14,
        6: 12,
        7: 12,
        8: 12,
        9: 12,
        10: 12,
        11: 14,
        12: 14,
        13: 14,
        14: 14,
        15: 18,
        16: 12,
        17: 12,
        18: 12,
        19: 14,
        20: 12,
        21: 12,
        22: 12,
    }
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def main() -> None:
    rows = build_rows()
    write_csv(rows)
    write_xlsx(rows)
    print(OUT_XLSX)
    print(OUT_CSV)


if __name__ == "__main__":
    main()
